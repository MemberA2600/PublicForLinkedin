import os
from tkinter.filedialog import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color, Font
from openpyxl.utils import get_column_letter
import re

def Copy_Months(workbook_output, path, months):
    for dirpath, dirnames, filenames in os.walk(path):
        for name in filenames:
            regex=re.findall(r'\d{4}', name.split("_")[0])[0]
            if regex!=None:
                year=int(regex)
                break


    for num in range (0, 12, 1):
        new = workbook_output.copy_worksheet(workbook_output.active)
        new["B3"]=year
        new["B4"]=months[num+1]
        new.title=months[num+1]
    del workbook_output["Lap"]
    return(workbook_output, year)

class Create_DataBase():
    def __init__(self, path, months):
        self.People={}

        for dirpath, dirnames, filenames in os.walk(path):
            for name in filenames:
                MonthSheet=load_workbook(dirpath+"/"+name)
                MonthName=name.split("_")[1].split(".")[0]
                self.People[MonthName]={}
                for person in MonthSheet.sheetnames:
                    self.People[MonthName][person]={}

                    for num in range(1, 32):
                        if MonthSheet[person][get_column_letter(num)+"17"].value==1 or MonthSheet[person][get_column_letter(num)+"17"].value=="1":
                            self.People[MonthName][person][num]=1
                        else:
                            self.People[MonthName][person][num]=0
                #print(self.People[MonthName])

def Add_Data_To_WorkSheet(workbook_output, database, months):
    from copy import copy
    from openpyxl.styles import PatternFill, Border, Side


    People=database.People
    for month in months.values():
        try:
            lenght=len(People[month])
        except:
            continue

        workbook_output[month].insert_rows(7, lenght-1)
        for num in range(0, lenght-2):
            workbook_output[month].merge_cells("A"+str(7+num)+":B"+str(7+num))

        for num in range(7, 9999999):
            workbook_output[month].merge_cells("A" + str(num) + ":B" + str(num))

            if workbook_output[month]["A"+str(num)].value=="Összesen:":

                workbook_output[month].merge_cells("A" + str(num+2) + ":B" + str(num+2))
                workbook_output[month].merge_cells("C" + str(num + 2) + ":D" + str(num + 2))
                sum=num
                break
            else:
                for cell in range(1, 34):
                    workbook_output[month][get_column_letter(cell)+str(num)]._style=copy(workbook_output[month][get_column_letter(cell)+str(num-1)]._style)


        #Buggy shit messes up emerged cells after insterting, LOL
        for num in range(0, lenght):
            workbook_output[month]["A"+str(6+num)].value=list(People[month].keys())[num]
            for day in People[month][list(People[month].keys())[num]]:
                workbook_output[month][get_column_letter(2+day)+str(num+6)].value=People[month][list(People[month].keys())[num]][day]

        all=0
        for row in range(3,34):
            just_the_row=0
            for col in range(6,sum):
                just_the_row+=int(workbook_output[month][get_column_letter(row)+str(col)].value)
            workbook_output[month][get_column_letter(row) + str(col+1)].value=just_the_row
            all+=just_the_row
        workbook_output[month]["C"+ str(col+3)].value=all

        if month=="Február":
            for num in range(6, sum + 1):

                workbook_output[month]["AF" + str(num)].fill = PatternFill(start_color='FF000000', end_color='FF000000',
                                                                           fill_type='solid')
                workbook_output[month]["AG" + str(num)].fill = PatternFill(start_color='FF000000', end_color='FF000000',
                                                                           fill_type='solid')
                workbook_output[month]["AE" + str(num)].fill = PatternFill(start_color='FF000000', end_color='FF000000',
                                                                           fill_type='solid')

                workbook_output[month]["AF" + str(num)].border = Border(left=Side(style='medium'),
                                                                        right=Side(style='medium'),
                                                                        top=Side(style='medium'),
                                                                        bottom=Side(style='medium'))

                workbook_output[month]["AE" + str(num)].border = Border(left=Side(style='medium'),
                                                                        right=Side(style='medium'),
                                                                        top=Side(style='medium'),
                                                                        bottom=Side(style='medium'))

                if (year%4==0 and year%100!=0) or year%400==0:
                    pass
                else:
                    workbook_output[month]["AD" + str(num)].fill = PatternFill(start_color='FF000000', end_color='FF000000',
                                                                               fill_type='solid')
                    workbook_output[month]["AD" + str(num)].border = Border(left=Side(style='medium'),
                                                                            right=Side(style='medium'),
                                                                            top=Side(style='medium'),
                                                                            bottom=Side(style='medium'))

        elif month in ["Január", "Március" , "Május", "Július", "Augusztus", "Október", "December"]:
            for num in range(6, sum+1):
                workbook_output[month]["AG"+str(num)].fill=PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
        else:
            for num in range(6, sum+1):
                workbook_output[month]["AF"+str(num)].fill=PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                workbook_output[month]["AG"+str(num)].fill=PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
                workbook_output[month]["AF"+str(num)].border=Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    return(workbook_output)

if __name__=="__main__":
    months={1: "Január", 2: "Február", 3: "Március", 4: "Április", 5: "Május", 6: "Június", 7: "Július", 8: "Augusztus", 9: "Szeptember", 10: "Október", 11: "November", 12: "December"}

    path = askdirectory(initialdir="*", title="Forrásmappa kiválasztása")
    workbook_output=load_workbook("template.xlsx")


    workbook_output, year = Copy_Months(workbook_output, path, months)
    database=Create_DataBase(path, months)

    workbook_output=Add_Data_To_WorkSheet(workbook_output, database, months)

    output_name=asksaveasfilename(initialdir="*", initialfile=str(year)+".xlsx" , title="Fájl kiválasztása", filetypes=(("excel 2007 fájlok", "*.xlsx"),))
    if output_name.endswith(".xlsx") == False:
        output_name=outputfile_name+".xlsx"
    workbook_output.save(output_name)