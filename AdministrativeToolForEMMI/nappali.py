import os
from tkinter import *
import re
from tkinter.filedialog import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color, Font
from openpyxl.utils import get_column_letter

import sys

def open_list(opendir, workbook_output):
    honapok = {"01": "Január", "02": "Február", "03": "Március", "04": "Április", "05": "Május", "06": "Június", "07": "Július", "08": "Augusztus", "09": "Szeptember", "10": "Október", "11": "November", "12": "December"}
    ped=messagebox.askyesno("Pedagógia", "Legyen pedagógiai segítségnyújtás?")
    gyogyped=messagebox.askyesno("Gyógyedagógia", "Legyen gyógypedagógiai segítségnyújtás?")
    gond=messagebox.askyesno("Gondozás", "Legyen (egészségügyi) gondozás?")
    name_n_days={}

    for dirpath, dirnames, filenames in os.walk(opendir):
        for name in filenames:
            if name.endswith(".xlsx") and "nappali" in name:
                #print("Megnyitás: "+ dirpath + "/" + name)
                workbook_day = load_workbook(dirpath+"/" +name)
                ev, honap, nap=get_date(workbook_day, name)
                workbook_output=add_date(workbook_output, ev, honap, honapok)

                names, name_n_days=create_sheets(workbook_day, name_n_days, nap)
                workbook_output=add_names(workbook_output, names)
                #print(name_n_days)
                workbook_output=add_values(workbook_day, workbook_output, ev, honap, nap, ped, gyogyped, gond)


    workbook_output=check_order(workbook_output)
    workbook_output=blackness(workbook_output, name_n_days)

    return (workbook_output, ev, honapok[honap])

def blackness(workbook_output, name_n_days):
    for sheet in workbook_output.sheetnames:
        active=workbook_output[sheet]
        for column in range(1, name_n_days[sheet]):
            for row in range(7, 27, 2):
                active[get_column_letter(column) + str(row)].value=0
                #active[get_column_letter(column)+str(row)].fill=PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')



    return workbook_output

def add_values(workbook_day, workbook_output, ev, honap, nap, ped, gyogyped, gond):
    nap=int(nap)
    honap=int(honap)
    ev=int(ev)

    for number in range(4,sys.maxsize):
        if workbook_day.active["C"+str(number)].value=="Összesen" or workbook_day.active["C"+str(number)].value==None:
            break
        active=workbook_output[workbook_day.active["C"+str(number)].value]
        if workbook_day.active["E"+str(number)].value==1 or workbook_day.active["F"+str(number)].value==1 or workbook_day.active["G"+str(number)].value==1 or workbook_day.active["H"+str(number)].value==1 or workbook_day.active["I"+str(number)].value==1 or workbook_day.active["J"+str(number)].value==1 or workbook_day.active["K"+str(number)].value==1 or workbook_day.active["N"+str(number)].value==1:
            active[str(get_column_letter(nap))+"17"].value=1
            if active.title=="Bobonkov Dusán":
                print(nap)
        else:
            active[str(get_column_letter(nap)) + "17"].value = 0
        if workbook_day.active["J"+str(number)].value==1:
            active[str(get_column_letter(nap))+"13"].value=1
        else:
            active[str(get_column_letter(nap)) + "13"].value = 0

        if workbook_day.active["H"+str(number)].value==1 or workbook_day.active["I"+str(number)].value==1:
            active[str(get_column_letter(nap))+"15"].value=1
        else:
            active[str(get_column_letter(nap)) + "15"].value = 0

        if gond==True and workbook_day.active["N"+str(number)].value==1:
            active[str(get_column_letter(nap)) + "19"].value = 1
        else:
            active[str(get_column_letter(nap)) + "19"].value = 0

        #print(active.title, workbook_day.active["C"+str(number)].value)
        for_random=0
        for charater in str(workbook_day.active["C"+str(number)].value):
            for_random=for_random+ord(charater)
        for_random*=nap
        for_random+=ev
        for_random+=honap
        for_random2=(bin(for_random))[2:]
        if len(for_random2)<6:
            for_random *= nap
            for_random2 = (bin(for_random))[2:]
        for_random=for_random2[::-1]

        if active[str(get_column_letter(nap)) + "15"].value==1:
            active[str(get_column_letter(nap)) + "7"].value = int(for_random[0])
            active[str(get_column_letter(nap)) + "9"].value = int(for_random[1])
            active[str(get_column_letter(nap)) + "11"].value = int(for_random[2])

            if ped==True:
                active[str(get_column_letter(nap)) + "21"].value = int(for_random[3])
            else:
                active[str(get_column_letter(nap)) + "21"].value = 0
            if gyogyped == True:
                active[str(get_column_letter(nap)) + "23"].value = int(for_random[4])
            else:
                active[str(get_column_letter(nap)) + "23"].value = 0

            active[str(get_column_letter(nap)) + "25"].value = int(for_random[5])
        else:
            active[str(get_column_letter(nap)) + "7"].value = 0
            active[str(get_column_letter(nap)) + "9"].value = 0
            active[str(get_column_letter(nap)) + "11"].value = 0
            active[str(get_column_letter(nap)) + "21"].value = 0
            active[str(get_column_letter(nap)) + "23"].value = 0
            active[str(get_column_letter(nap)) + "25"].value = 0
    return workbook_output

def check_order(workbook_output):
    names=workbook_output.sheetnames
    sorted_names=[]
    for name in names:
        sorted_names.append(name)
    sorted_names.sort()

    counter=0
    while counter<len(names):
        moving = workbook_output[sorted_names[counter]]
        workbook_output.move_sheet(moving, (len(names)-workbook_output.index(moving)))
        counter+=1

    return workbook_output

def add_names(workbook_output, names):
    for name in names:

        if name in workbook_output.sheetnames:
            continue
        if name!=None:

            new=workbook_output.copy_worksheet(workbook_output.active)
            new["A2"].value="Ellátott neve: "+name
            new.title=name

    try:
        #workbook_output.remove_sheet(workbook_output.get_sheet_by_name("template"))
        del workbook_output["template"]
    except:
        pass
    return(workbook_output)

def add_date(workbook_output, ev, honap, honapok):
    workbook_output.active["E3"].value=int(ev)
    workbook_output.active["I3"].value=honapok[honap]
    for number in range(7, 27, 2):
        workbook_output.active["AF" + str(number)].font=Font(color="FFFFFF")

    if honap in ["02", "04", "06", "09", "11"]:

        for number in range(7, 27, 2):

            workbook_output.active["AE" + str(number)].value = 0
            workbook_output.active["AE"+str(number)].fill=PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
            if honap=="02":
                workbook_output.active["AD" + str(number)].value = 0
                workbook_output.active["AD" + str(number)].fill = PatternFill(start_color='FF000000',
                                                                              end_color='FF000000', fill_type='solid')
                if int(ev)%4==0 and int(ev)%100!=0:
                    continue
                if int(ev) % 400 == 0:
                    continue

                workbook_output.active["AC" + str(number)].value = 0
                workbook_output.active["AC" + str(number)].fill = PatternFill(start_color='FF000000',
                                                                      end_color='FF000000',
                                                                      fill_type='solid')
    return(workbook_output)


def create_sheets(workbook_day, name_n_days, nap):
    names=[]
    for number in range(4,sys.maxsize):
        if workbook_day.active["C"+str(number)].value!="Összesen" and workbook_day.active["C"+str(number)].value!=None:
            names.append((workbook_day.active["C"+str(number)].value))
            if workbook_day.active["C"+str(number)].value in name_n_days:
                continue
            else:
                name_n_days.setdefault(workbook_day.active["C"+str(number)].value, int(nap))

        else:
            break
    return names, name_n_days

def get_date(workbook_day, name):
    date = None
    try:
        for number in range(0, len(workbook_day.sheetnames) + 1):
            workbook_day.active = number

            try:

                trial = re.findall(r'\d\d\d\d-\d\d-\d\d', str(workbook_day.active["B4"].value))[0]
                date = trial
                ev = date.split("-")[0]
                honap = date.split("-")[1]
                nap = date.split("-")[2]
                break
            except:
                continue
    except:
        try:
            date = re.findall(r'\d\d\d\d\.\d\d\.\d\d', name)[0]
            ev = date.split(".")[0]
            honap = date.split(".")[1]
            nap = date.split(".")[2]
        except:
            ev=None
            honap=None
            nap=None
    return(ev,honap,nap)

if  __name__=="__main__":
    workbook_output=load_workbook("template.xlsx")
    opendir=askdirectory(initialdir="*", title="Forrásmappa kiválasztása")
    workbook_output, ev, honap=open_list(opendir, workbook_output)


    outputfile = asksaveasfile(initialdir="*", initialfile=ev+"_"+honap+".xlsx" , title="Fájl kiválasztása", filetypes=(("excel 2007 fájlok", "*.xlsx"),))
    outputfile_name=outputfile.name
    outputfile.close()
    if outputfile_name.endswith(".xlsx") == False:
        outputfile_name=outputfile_name+".xlsx"
    workbook_output.save(outputfile_name)

